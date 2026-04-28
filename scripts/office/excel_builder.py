"""Generate formatted Excel reports from a JSON report spec using openpyxl.

Usage:
    python excel_builder.py --spec report_spec.json --output report.xlsx

The JSON spec format:
{
  "title": "Sprint 42 Tracker",
  "type": "sprint-tracker",
  "sheets": [
    {
      "name": "Backlog",
      "headers": ["ID", "Title", "Type", "Assignee", "Status", "Points", "Sprint"],
      "rows": [
        [1042, "Migrate auth to Azure AD", "Story", "Alice", "Done", 8, "Sprint 42"]
      ],
      "status_column": 4,
      "freeze_row": 1
    }
  ]
}

Conventions:
- Header rows: bold, light blue fill
- Status columns: green (Done/Closed), amber (Active/In Progress), red (Blocked)
- Auto-fit column widths
- Auto-filter on header row
- Freeze top row
"""

import json
from pathlib import Path

import typer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.console import Console

app = typer.Typer()
console = Console()

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)

STATUS_FILLS = {
    "done": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "active": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "blocked": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

STATUS_MAP = {
    "done": {"done", "closed", "resolved", "completed"},
    "active": {"active", "in progress", "in review", "committed"},
    "blocked": {"blocked", "impediment"},
}


def get_status_category(value: str) -> str | None:
    v = str(value).lower().strip()
    for category, states in STATUS_MAP.items():
        if v in states:
            return category
    return None


def auto_fit_columns(ws: "Worksheet") -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def write_sheet(wb: Workbook, sheet_spec: dict) -> None:
    name = sheet_spec.get("name", "Sheet")
    headers = sheet_spec.get("headers", [])
    rows = sheet_spec.get("rows", [])
    status_col_idx = sheet_spec.get("status_column")  # 0-based column index of status column
    freeze = sheet_spec.get("freeze_row", 1)

    ws = wb.create_sheet(title=name)

    if headers:
        for col_i, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_i, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = ws.cell(row=freeze + 1, column=1)

    for row_i, row in enumerate(rows, start=2):
        for col_i, value in enumerate(row, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            if status_col_idx is not None and col_i == status_col_idx + 1:
                cat = get_status_category(str(value))
                if cat:
                    cell.fill = STATUS_FILLS[cat]

    auto_fit_columns(ws)


def build_report(spec: dict, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    sheets = spec.get("sheets", [])
    if not sheets:
        console.print("[yellow]No sheets defined in spec.[/yellow]")

    for sheet_spec in sheets:
        write_sheet(wb, sheet_spec)

    wb.save(str(output_path))


@app.command()
def main(
    spec: Path = typer.Option(..., help="Path to JSON report spec"),
    output: Path = typer.Option(..., help="Output .xlsx file path"),
) -> None:
    """Generate a formatted Excel report from a JSON spec."""
    if not spec.exists():
        console.print(f"[red]Spec file not found:[/red] {spec}")
        raise typer.Exit(1)

    report_spec = json.loads(spec.read_text())
    if not isinstance(report_spec, dict):
        console.print("[red]Spec must be a JSON object.[/red]")
        raise typer.Exit(1)

    title = report_spec.get("title", "Report")
    sheet_count = len(report_spec.get("sheets", []))
    console.print(f"Building '{title}' with {sheet_count} sheet(s)...")
    build_report(report_spec, output)
    console.print(f"[green]✓ Report saved to {output}[/green]")


if __name__ == "__main__":
    app()
